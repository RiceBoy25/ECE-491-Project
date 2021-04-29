# -*- coding: utf-8 -*-
"""
Created on Tue Apr 20 19:23:52 2021

@author: Eric
"""

import pandas as pd
import xlrd
import openpyxl

#https://www.geeksforgeeks.org/python-reading-excel-file-using-openpyxl-module/
path = ""


def fillStock(stockArray, stockData):
    i = 2
    #numSales = len(stockData)
    numSales = stockData.max_row
    print("number of sales: ", numSales)
    while i <= numSales:
        stockArray.append(stockData.cell(row = i, column = 2).value)
        i = i + 1
    stockArray.reverse()
    print(stockArray)
    return stockArray

def predict(stockArray, weights, index):
    i = 0
    total = 0
    while i < len(weights):
        total = total + (stockArray[index + i] * weights[i])
        i = i + 1
    print(total)
    return total
#predict(dataStock, weights, 1)

def errorRate(predicted, actual):
    return actual - predicted

def denominator(stockArray, size, index):
    i = 0
    total = 0
    while i < size:
        total = total + (stockArray[index + i]**2)
        i = i + 1
    return total

def weightDiff(stockArray, size, index, dom, error):
    i = 0
    weightSub = [0] * size
    fraction = (-1 * error) / dom
    while i < size:
        weightSub[i] = fraction * stockArray[index + i]
        i = i + 1
    return weightSub

def newWeight(old, new, size):
    i = 0
    finalWeight = [0] * size
    while i < size:
        finalWeight[i] = old[i] - new[i]
        i = i + 1
    return finalWeight

def master(stockArray, stockData, size):
    stockArray = fillStock(stockArray, stockData)
    #size refers to the number of data entries being fed into the LMS
    #algorithm at a single time
    weights = [1] * size
    print("Weights: ", weights)
    print("Stock:\n", stockArray)
    index = 0
    numSales = stockData.max_row
    while index <= (numSales - size):
        predicted = predict(stockArray, weights, index)
        error = errorRate(predicted, stockArray[index + size])
        dom = denominator(stockArray, size, index)
        weightSub = weightDiff(stockArray, size, index, dom, error)
        weights = newWeight(weights, weightSub, size)
        index = index + 1
        print(index, ": ", "Actual: ", stockArray[index], " Predicted: ", predicted)
        #print("Accuracy: ", (predicted / stockArray[index]), "\n")
        print("Accuracy: ", abs(1 - (abs(predicted - stockArray[index])/stockArray[index])), "\n")

print("Starting Program...")
while 1:
    userChoice = input("Enter stock or crypto name, 0 to exit\n")
    if userChoice == "Google" or userChoice == "Apple" or userChoice == "Doge" :
        quoteLength = input("Enter quote length in months: 1m, 3m, or 6m\n")
        userFilter = input("Enter filter size\n")
        path = userChoice + quoteLength + ".xlsx" 
        data = openpyxl.load_workbook(path)
        data_obj = data.active
        stock = data_obj.cell(row = 2, column = 2)
        dataStock = []
        fillStock(dataStock, data_obj)
        size = int(userFilter) #filter size
        weights = [1] * size
        print("weights: ", weights[1])
        master(dataStock, data_obj, 2)
    elif userChoice == "0":
        print("Ending program...")
        break
    else:
        print("Invalid entry... Try again")
        







"""
Code block tools...................................
"""

"""
fillStock(dataStock, data_obj)
size = 2 #filter size
weights = [1] * size
print("weights: ", weights[1])

master(dataStock, data_obj, 2)
"""

    